"""
pyCoreGage.reporter
===================
build_reports() — merges findings with feedback, tracks status changes,
and writes six role-based Excel reports.

Reports written
---------------
DM_issues.xlsx      — Data Management findings
MW_issues.xlsx      — Medical Writing findings
SDTM_issues.xlsx    — SDTM programmer findings
ADAM_issues.xlsx    — ADaM programmer findings
all_open.xlsx       — All open / queried findings
all_closed.xlsx     — All closed findings
"""

from __future__ import annotations

import logging
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .state import CoreGageConfig, CoreGageState
from .utils import cleanup_text

logger = logging.getLogger("pyCoreGage")

# ---------------------------------------------------------------------------
# Date parsing helpers
# ---------------------------------------------------------------------------

def _parse_date_col(series: pd.Series) -> pd.Series:
    """
    Parse a mixed-format date column into datetime.date objects.

    Handles:
    - Excel serial numbers  (numeric strings like "45123")
    - ISO strings           ("2024-01-15")
    - ddMMMYYYY format      ("15Jan2024")
    """
    result = pd.Series([None] * len(series), index=series.index, dtype=object)

    for i, val in series.items():
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        if isinstance(val, date):
            result[i] = val
            continue
        s = str(val).strip()
        # Try Excel serial
        try:
            n = float(s)
            result[i] = (pd.Timestamp("1899-12-30") + pd.Timedelta(days=n)).date()
            continue
        except (ValueError, OverflowError):
            pass
        # Try ISO
        try:
            result[i] = pd.to_datetime(s, format="%Y-%m-%d").date()
            continue
        except (ValueError, TypeError):
            pass
        # Try ddMMMYYYY
        try:
            result[i] = datetime.strptime(s, "%d%b%Y").date()
            continue
        except (ValueError, TypeError):
            pass
    return result


# ---------------------------------------------------------------------------
# Import previously saved reports
# ---------------------------------------------------------------------------

_SAVED_COL_MAP = {
    "CHECK ID":     "id",
    "SUBJECT ID":   "subj_id",
    "VISIT ID":     "vis_id",
    "DESCRIPTION":  "desrp",
    "FIND DATE":    "find_dt",
    "STATUS":       "status",
    "ANALYST NOTE": "analyst_note",
    "ANALYST ID":   "analyst_id",
    "REVIEW NOTE":  "review_note",
    "REVIEWER ID":  "reviewer_id",
}

_SAVED_NEEDED = [
    "id", "subj_id", "vis_id", "desrp", "find_dt",
    "status", "analyst_note", "analyst_id",
]


def _import_saved(reports_dir: str, fname: str) -> Optional[pd.DataFrame]:
    """Read a previously written report's Details sheet."""
    path = os.path.join(reports_dir, fname)
    if not os.path.isfile(path):
        return None
    logger.info("  Importing saved issues: %s", fname)
    try:
        df = pd.read_excel(path, sheet_name="Details", skiprows=2, dtype=str)
    except Exception as exc:
        logger.warning("  WARNING: cannot read %s -- %s", fname, exc)
        return None

    if df is None or df.empty:
        return None

    df = df.rename(columns=_SAVED_COL_MAP)

    for col in _SAVED_NEEDED:
        if col not in df.columns:
            df[col] = pd.NA

    df = df[_SAVED_NEEDED].copy()
    df = df.dropna(subset=["find_dt"])

    # Filter placeholder analyst notes
    if "analyst_note" in df.columns:
        placeholder = df["analyst_note"].astype(str).str.upper().str.strip() == "THIS IS EMPTY"
        df = df[~placeholder]

    if df.empty:
        return None

    df["vis_id"]  = pd.to_numeric(df["vis_id"], errors="coerce")
    df["find_dt"] = _parse_date_col(df["find_dt"])
    df["desrp"]       = df["desrp"].astype(str).apply(cleanup_text)
    df["analyst_note"] = df["analyst_note"].astype(str).apply(cleanup_text)
    df["dup_id"]  = df["desrp"].astype(str).apply(lambda x: re.sub(r"\s", "", x))

    out = df[["id", "subj_id", "vis_id", "desrp", "dup_id",
              "find_dt", "status", "analyst_note", "analyst_id"]].copy()
    logger.info("    -> %d saved issue(s) loaded from %s", len(out), fname)
    return out


# ---------------------------------------------------------------------------
# Feedback reader
# ---------------------------------------------------------------------------

_FB_COL_MAP = {
    "CHECK ID":     "id",
    "SUBJECT ID":   "subj_id",
    "VISIT ID":     "vis_id",
    "DESCRIPTION":  "desrp",
    "STATUS":       "fb_status",
    "REVIEW NOTE":  "review_note",
    "REVIEWER ID":  "reviewer_id",
    "ANALYST NOTE": "fb_analyst_note",
    "ANALYST ID":   "fb_analyst_id",
}


def _read_feedback(feedback_dir: str, role: str) -> Optional[pd.DataFrame]:
    """
    Read the most-recently-modified feedback file for a given role.

    Returns None if no feedback file exists.
    """
    role_dir = os.path.join(feedback_dir, role)
    if not os.path.isdir(role_dir):
        return None

    xlsx_files = list(Path(role_dir).glob("*.xlsx"))
    if not xlsx_files:
        return None

    # Most recently modified file
    latest = max(xlsx_files, key=lambda p: p.stat().st_mtime)
    logger.info("  Reading feedback (%s): %s", role, latest.name)

    try:
        df = pd.read_excel(str(latest), sheet_name="Details", skiprows=2, dtype=str)
    except Exception as exc:
        logger.warning("  WARNING: cannot read feedback %s: %s", latest.name, exc)
        return None

    if df is None or df.empty:
        return None

    df = df.rename(columns=_FB_COL_MAP)

    needed = ["id", "subj_id", "vis_id", "desrp",
              "fb_status", "review_note", "reviewer_id",
              "fb_analyst_note", "fb_analyst_id"]
    for col in needed:
        if col not in df.columns:
            df[col] = pd.NA

    df = df[needed].copy()
    df["vis_id"] = pd.to_numeric(df["vis_id"], errors="coerce")
    df["desrp"]  = df["desrp"].astype(str).apply(cleanup_text)
    df["dup_id"] = df["desrp"].apply(lambda x: re.sub(r"\s", "", x))
    return df


# ---------------------------------------------------------------------------
# NA sentinel for merge key (pandas merge does not match NaN == NaN)
# ---------------------------------------------------------------------------

_VIS_SENTINEL = -1.0


def _add_vis_sentinel(df: pd.DataFrame) -> pd.DataFrame:
    if "vis_id" in df.columns:
        df = df.copy()
        df["vis_id"] = df["vis_id"].fillna(_VIS_SENTINEL)
    return df


def _drop_vis_sentinel(df: pd.DataFrame) -> pd.DataFrame:
    if "vis_id" in df.columns:
        df = df.copy()
        df["vis_id"] = df["vis_id"].replace(_VIS_SENTINEL, float("nan"))
    return df


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

_GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
_AMBER_FILL  = PatternFill("solid", fgColor="FFEB9C")
_RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
_STRIPE_FILL = PatternFill("solid", fgColor="EAF1DD")
_HEADER_FONT = Font(bold=True)
_WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")


def _auto_width(ws, max_width: int = 60) -> None:
    """Set approximate column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_report(
    out_name: str,
    reports_dir: str,
    sum_df: pd.DataFrame,
    det_df: pd.DataFrame,
    title: str,
    run_label: str,
) -> None:
    """Write a single report workbook with Summary and Details sheets."""
    outpath = os.path.join(reports_dir, f"{out_name}.xlsx")
    logger.info("  Writing: %s", os.path.basename(outpath))

    wb = Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    if sum_df.empty:
        sum_df = pd.DataFrame({
            "Check Description": ["No issues found"],
            "Check ID":          ["N/A"],
            "New":               [0],
            "Open":              [0],
            "Closed":            [0],
        })
    else:
        s_cols = [c for c in ["description", "id", "nu", "n_open", "n_closed"]
                  if c in sum_df.columns]
        sum_df = sum_df[s_cols].copy()
        col_rename = {
            "description": "Check Description",
            "id":          "Check ID",
            "nu":          "New",
            "n_open":      "Open",
            "n_closed":    "Closed",
        }
        sum_df = sum_df.rename(columns=col_rename)

    ws_sum["A1"] = f"pyCoreGage Report — {title}"
    ws_sum["A1"].font = Font(bold=True, size=13)
    ws_sum["A2"] = run_label

    # Write headers at row 4
    header_row = 4
    for ci, col_name in enumerate(sum_df.columns, start=1):
        cell = ws_sum.cell(row=header_row, column=ci, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")

    # Write data rows + conditional colour on Open / Closed columns
    open_col_idx   = list(sum_df.columns).index("Open")   + 1 if "Open"   in sum_df.columns else None
    closed_col_idx = list(sum_df.columns).index("Closed") + 1 if "Closed" in sum_df.columns else None

    for ri, (_, row) in enumerate(sum_df.iterrows(), start=header_row + 1):
        for ci, val in enumerate(row, start=1):
            ws_sum.cell(row=ri, column=ci, value=val)
        # Colour Open column
        if open_col_idx:
            n_open = row.get("Open", 0) or 0
            fill = _GREEN_FILL if int(n_open) == 0 else _AMBER_FILL
            ws_sum.cell(row=ri, column=open_col_idx).fill = fill
        # Colour Closed column
        if closed_col_idx:
            n_closed = row.get("Closed", 0) or 0
            fill = _RED_FILL if int(n_closed) == 0 else _AMBER_FILL
            ws_sum.cell(row=ri, column=closed_col_idx).fill = fill

    _auto_width(ws_sum)

    # ── Details sheet ────────────────────────────────────────────────────────
    ws_det = wb.create_sheet("Details")

    _DET_COLS = [
        "id", "subj_id", "vis_id", "desrp", "find_dt",
        "status", "analyst_note", "analyst_id", "review_note", "reviewer_id",
    ]
    _DET_LABELS = [
        "CHECK ID", "SUBJECT ID", "VISIT ID", "DESCRIPTION", "FIND DATE",
        "STATUS", "ANALYST NOTE", "ANALYST ID", "REVIEW NOTE", "REVIEWER ID",
    ]

    if det_df.empty:
        det_df = pd.DataFrame([{
            "id":           "N/A",
            "subj_id":      "N/A",
            "vis_id":       "",
            "desrp":        "No issues found",
            "find_dt":      date.today(),
            "status":       "N/A",
            "analyst_note": "This is empty",
            "analyst_id":   "",
            "review_note":  "",
            "reviewer_id":  "",
        }])

    for col in _DET_COLS:
        if col not in det_df.columns:
            det_df[col] = ""

    dout = det_df[[c for c in _DET_COLS if c in det_df.columns]].copy()

    ws_det["A1"] = f"All {title} Issues"
    ws_det["A1"].font = Font(bold=True, size=13)
    ws_det["A2"] = run_label

    header_row_det = 3
    for ci, label in enumerate(_DET_LABELS[:len(dout.columns)], start=1):
        cell = ws_det.cell(row=header_row_det, column=ci, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="203864")

    for ri, (_, row) in enumerate(dout.iterrows(), start=header_row_det + 1):
        for ci, val in enumerate(row, start=1):
            cell = ws_det.cell(row=ri, column=ci, value=val)
            cell.alignment = _WRAP_ALIGN
        # Alternating row stripe
        if (ri - header_row_det) % 2 == 0:
            for ci in range(1, len(dout.columns) + 1):
                ws_det.cell(row=ri, column=ci).fill = _STRIPE_FILL

    _auto_width(ws_det)

    wb.save(outpath)


# ---------------------------------------------------------------------------
# Main public function
# ---------------------------------------------------------------------------

def build_reports(cfg: CoreGageConfig, state: CoreGageState) -> None:
    """
    Consolidate findings and write role-based Excel reports.

    Merges new findings with previously saved issues, incorporates reviewer
    feedback from role-separated feedback folders, applies smart status
    management rules, and writes six Excel reports to ``cfg.reports``.

    Feedback loop
    -------------
    After distributing reports, reviewers place updated files in the
    appropriate subfolder under ``cfg.feedback``:

    - ``feedback/DM/``   — Data Management reviewer
    - ``feedback/MW/``   — Medical Writing reviewer
    - ``feedback/SDTM/`` — SDTM programmer
    - ``feedback/ADAM/`` — ADaM programmer

    On the next run, ``build_reports()`` reads the most recent file from
    each folder, merges ``analyst_note``, ``review_note``, and status
    updates back into the findings, and overwrites the reports with the
    merged version.

    Smart status management
    -----------------------
    - Findings that disappear from the data are automatically closed.
    - Findings closed by a reviewer remain closed on future runs.
    - Findings that re-appear after analyst closure are re-opened.
    - Duplicate auto-tags are never appended twice.

    Parameters
    ----------
    cfg : CoreGageConfig
        Project configuration.
    state : CoreGageState
        State returned by :func:`run_checks`.

    Returns
    -------
    None
        Reports are written to ``cfg.reports``.
    """
    logger.info(">> [reporter] Starting consolidation ...")

    sdate = state.session.get("sdate", date.today().strftime("%d%b%Y"))
    stime = state.session.get("stime", datetime.now().strftime("%H:%M"))
    run_label = f"Last run: {sdate} {stime}"

    os.makedirs(cfg.reports, exist_ok=True)

    # ── Load previously saved issues ─────────────────────────────────────────
    old_open   = _import_saved(cfg.reports, "all_open.xlsx")
    old_closed = _import_saved(cfg.reports, "all_closed.xlsx")

    def _empty_old() -> pd.DataFrame:
        return pd.DataFrame({
            "id": pd.Series(dtype=str), "subj_id": pd.Series(dtype=str),
            "vis_id": pd.Series(dtype=float), "desrp": pd.Series(dtype=str),
            "dup_id": pd.Series(dtype=str), "find_dt": pd.Series(dtype=object),
            "status": pd.Series(dtype=str), "analyst_note": pd.Series(dtype=str),
            "analyst_id": pd.Series(dtype=str),
        })

    parts = [df for df in [old_open, old_closed] if df is not None]
    olddata = pd.concat(parts, ignore_index=True) if parts else _empty_old()

    has_new = not state.issues.empty
    has_old = not olddata.empty

    # ── Merge new findings with old ──────────────────────────────────────────
    if not has_new and not has_old:
        logger.info("  No findings to report. Writing empty reports.")
        issuelist = pd.DataFrame(columns=[
            "id", "subj_id", "vis_id", "desrp", "dup_id", "find_dt",
            "status", "analyst_note", "analyst_id",
            "review_note", "reviewer_id",
            "category", "subcategory", "description",
            "dm_report", "mw_report", "sdtm_report", "adam_report",
        ])
    else:
        if has_new:
            newf = state.issues.copy()
            newf = newf.rename(columns={"description": "desrp"})
            newf = newf[["id", "subj_id", "vis_id", "desrp"]].drop_duplicates()
            newf["dup_id"] = newf["desrp"].astype(str).apply(
                lambda x: re.sub(r"\s", "", x)
            )
            newf["_new"] = True
            newf = _add_vis_sentinel(newf)
        else:
            newf = pd.DataFrame(
                columns=["id", "subj_id", "vis_id", "desrp", "dup_id", "_new"]
            )

        if has_old:
            olddata["_old"] = True
            olddata = _add_vis_sentinel(olddata)

        merge_keys = ["id", "subj_id", "vis_id", "dup_id"]

        if has_new and has_old:
            merged = pd.merge(newf, olddata, on=merge_keys, how="outer")
            merged["_new"] = merged["_new"].fillna(False)
            merged["_old"] = merged["_old"].fillna(False)

            # Resolve duplicate desrp columns
            if "desrp_x" in merged.columns and "desrp_y" in merged.columns:
                merged["desrp"] = merged["desrp_x"].combine_first(merged["desrp_y"])
                merged = merged.drop(columns=["desrp_x", "desrp_y"])
        elif has_new:
            merged = newf.copy()
            merged["_old"]         = False
            merged["find_dt"]      = None
            merged["status"]       = pd.NA
            merged["analyst_note"] = pd.NA
            merged["analyst_id"]   = pd.NA
        else:
            merged = olddata.copy()
            merged["_new"] = False

        # ── Assign finding dates ──────────────────────────────────────────
        today = date.today()
        new_and_not_old = merged["_new"] & ~merged["_old"]

        if "find_dt" not in merged.columns:
            merged["find_dt"] = None

        merged.loc[new_and_not_old & merged["find_dt"].isna(), "find_dt"] = today

        # ── Assign status ─────────────────────────────────────────────────
        if "status" not in merged.columns:
            merged["status"] = pd.NA

        # New finding not previously seen → Open
        merged.loc[new_and_not_old & merged["status"].isna(), "status"] = "open"

        # Disappeared findings → auto-close
        old_only = merged["_old"] & ~merged["_new"]
        if "analyst_note" not in merged.columns:
            merged["analyst_note"] = ""
        merged["analyst_note"] = merged["analyst_note"].fillna("").astype(str)

        not_already_closed = merged["status"].astype(str).str.lower().str.strip() != "closed"
        auto_close_mask = old_only & not_already_closed
        merged.loc[auto_close_mask, "status"] = "closed"
        already_tagged = merged["analyst_note"].str.contains(
            r"\[auto-closed — finding no longer present\]", regex=False, na=False
        )
        merged.loc[auto_close_mask & ~already_tagged, "analyst_note"] = (
            "[auto-closed — finding no longer present] "
            + merged.loc[auto_close_mask & ~already_tagged, "analyst_note"]
        ).str.strip()

        # Re-appeared after analyst-closed → re-open
        analyst_closed = merged["analyst_note"].str.contains(
            "closed by analyst", case=False, na=False
        )
        not_reviewer_closed = ~merged["analyst_note"].str.contains(
            "closed by reviewer", case=False, na=False
        )
        re_appeared = merged["_new"] & merged["_old"] & analyst_closed & not_reviewer_closed
        merged.loc[re_appeared, "status"] = "open"
        already_reappeared = merged["analyst_note"].str.contains(
            "[Was closed but re-appeared]", regex=False, na=False
        )
        merged.loc[re_appeared & ~already_reappeared, "analyst_note"] = (
            "[Was closed but re-appeared] "
            + merged.loc[re_appeared & ~already_reappeared, "analyst_note"]
        ).str.strip()

        # Fill remaining status
        merged["status"] = merged["status"].fillna("open")

        # Fill notes
        for col in ["analyst_note", "analyst_id", "review_note", "reviewer_id"]:
            if col not in merged.columns:
                merged[col] = ""
            merged[col] = merged[col].fillna("").astype(str)

        merged = _drop_vis_sentinel(merged)
        issuelist = merged.copy()
        if "dup_id" in issuelist.columns:
            issuelist = issuelist.drop(columns=["dup_id"])

        # ── Merge feedback from each role ─────────────────────────────────
        for role in ["DM", "MW", "SDTM", "ADAM"]:
            fb = _read_feedback(cfg.feedback, role)
            if fb is None or fb.empty:
                continue

            fb_keys = ["id", "subj_id", "vis_id", "dup_id"]
            issuelist["dup_id"] = issuelist["desrp"].astype(str).apply(
                lambda x: re.sub(r"\s", "", x)
            )
            issuelist = _add_vis_sentinel(issuelist)
            fb = _add_vis_sentinel(fb)

            # Drop any fb column whose suffixed name already exists on the left side.
            # pandas raises MergeError when suffixes would produce a column that
            # already exists (e.g. "desrp.fb" when issuelist already has "desrp.fb").
            fb_extra = [c for c in fb.columns if c not in fb_keys]
            left_cols = set(issuelist.columns)
            safe_fb_extra = [
                c for c in fb_extra
                if not (c in left_cols and (c + ".fb") in left_cols)
            ]
            fb_right = fb[fb_keys + safe_fb_extra].copy()

            merged_fb = pd.merge(
                issuelist, fb_right,
                on=fb_keys, how="left", suffixes=("", ".fb"),
            )

            # Apply reviewer status
            if "fb_status" in merged_fb.columns:
                reviewer_closed = (
                    merged_fb["fb_status"].astype(str).str.lower().str.strip() == "closed"
                )
                already_rc = merged_fb["analyst_note"].str.contains(
                    "[closed by reviewer]", regex=False, na=False
                )
                merged_fb.loc[reviewer_closed, "status"] = "closed"
                merged_fb.loc[reviewer_closed & ~already_rc, "analyst_note"] = (
                    "[closed by reviewer] "
                    + merged_fb.loc[reviewer_closed & ~already_rc, "analyst_note"]
                ).str.strip()
                merged_fb = merged_fb.drop(columns=["fb_status"])

            # Merge review_note
            if "review_note.fb" in merged_fb.columns:
                merged_fb["review_note"] = merged_fb["review_note.fb"].where(
                    merged_fb["review_note.fb"].notna()
                    & (merged_fb["review_note.fb"].astype(str).str.strip() != ""),
                    other=merged_fb.get("review_note", ""),
                )
                merged_fb = merged_fb.drop(columns=["review_note.fb"])

            if "reviewer_id.fb" in merged_fb.columns:
                merged_fb["reviewer_id"] = merged_fb["reviewer_id.fb"].where(
                    merged_fb["reviewer_id.fb"].notna()
                    & (merged_fb["reviewer_id.fb"].astype(str).str.strip() != ""),
                    other=merged_fb.get("reviewer_id", ""),
                )
                merged_fb = merged_fb.drop(columns=["reviewer_id.fb"])

            if "fb_analyst_note" in merged_fb.columns:
                merged_fb["analyst_note"] = merged_fb["fb_analyst_note"].where(
                    merged_fb["fb_analyst_note"].notna()
                    & (merged_fb["fb_analyst_note"].astype(str).str.strip() != ""),
                    other=merged_fb["analyst_note"],
                )
                merged_fb = merged_fb.drop(columns=["fb_analyst_note"])

            if "fb_analyst_id" in merged_fb.columns:
                merged_fb["analyst_id"] = merged_fb["fb_analyst_id"].where(
                    merged_fb["fb_analyst_id"].notna()
                    & (merged_fb["fb_analyst_id"].astype(str).str.strip() != ""),
                    other=merged_fb.get("analyst_id", ""),
                )
                merged_fb = merged_fb.drop(columns=["fb_analyst_id"])

            issuelist = _drop_vis_sentinel(merged_fb)
            if "dup_id" in issuelist.columns:
                issuelist = issuelist.drop(columns=["dup_id"])

        # Drop helper columns
        for col in ["_new", "_old"]:
            if col in issuelist.columns:
                issuelist = issuelist.drop(columns=[col])

    # ── Attach registry metadata ──────────────────────────────────────────────
    reg = state.rule_registry[
        [c for c in ["id", "description", "category", "subcategory",
                      "dm_report", "mw_report", "sdtm_report", "adam_report"]
         if c in state.rule_registry.columns]
    ].drop_duplicates(subset=["id"])

    if not issuelist.empty:
        issuelist = pd.merge(issuelist, reg, on="id", how="left")
    else:
        for col in reg.columns:
            if col not in issuelist.columns:
                issuelist[col] = pd.NA

    # ── Summary statistics ────────────────────────────────────────────────────
    status_col = issuelist["status"].astype(str).str.lower().str.strip() if not issuelist.empty else pd.Series(dtype=str)
    n_open_tot   = (status_col != "closed").sum() if not issuelist.empty else 0
    n_closed_tot = (status_col == "closed").sum() if not issuelist.empty else 0
    n_queried    = (status_col == "queried").sum() if not issuelist.empty else 0
    n_analyst = 0
    n_reviewer = 0
    if not issuelist.empty and "analyst_note" in issuelist.columns:
        notes = issuelist["analyst_note"].astype(str).str.strip()
        n_analyst  = ((notes != "") & ~notes.str.startswith("[")).sum()
    if not issuelist.empty and "review_note" in issuelist.columns:
        rnotes = issuelist["review_note"].astype(str).str.strip()
        n_reviewer = (rnotes != "").sum()

    logger.info("  -------------------------------------------------------")
    logger.info("  Feedback summary:")
    logger.info("    Notes  : analyst notes: %d  |  reviewer notes: %d",
                n_analyst, n_reviewer)
    logger.info("    Status : open: %d  |  queried: %d  |  closed: %d",
                n_open_tot, n_queried, n_closed_tot)
    logger.info("  -------------------------------------------------------")

    # ── Build head_sum ────────────────────────────────────────────────────────
    sl = state.summary_log[["headlink", "nu"]].rename(columns={"headlink": "id"})

    def _count_by(df: pd.DataFrame, col_name: str) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame({"id": pd.Series(dtype=str), col_name: pd.Series(dtype=int)})
        counts = df.groupby("id").size().reset_index(name=col_name)
        return counts

    open_iss   = issuelist[status_col != "closed"] if not issuelist.empty else issuelist
    closed_iss = issuelist[status_col == "closed"] if not issuelist.empty else issuelist

    n_open_by   = _count_by(open_iss,   "n_open")
    n_closed_by = _count_by(closed_iss, "n_closed")

    head_sum = reg.copy()
    head_sum = pd.merge(head_sum, sl, on="id", how="left")
    head_sum = pd.merge(head_sum, n_open_by, on="id", how="left")
    head_sum = pd.merge(head_sum, n_closed_by, on="id", how="left")
    head_sum["nu"]       = head_sum["nu"].fillna(0).astype(int)
    head_sum["n_open"]   = head_sum["n_open"].fillna(0).astype(int)
    head_sum["n_closed"] = head_sum["n_closed"].fillna(0).astype(int)

    # ── Write all six reports ─────────────────────────────────────────────────
    def _rpt_mask(col: str) -> pd.Series:
        if col not in head_sum.columns:
            return pd.Series(False, index=head_sum.index)
        return head_sum[col].astype(str).str.upper().str.startswith("Y")

    def _rpt_det(df: pd.DataFrame, col: str) -> pd.DataFrame:
        if df.empty or col not in df.columns:
            return df
        return df[df[col].astype(str).str.upper().str.startswith("Y")]

    _write_report("DM_issues",   cfg.reports,
                  head_sum[_rpt_mask("dm_report")],
                  _rpt_det(open_iss, "dm_report"),   "DM",   run_label)
    _write_report("MW_issues",   cfg.reports,
                  head_sum[_rpt_mask("mw_report")],
                  _rpt_det(open_iss, "mw_report"),   "MW",   run_label)
    _write_report("SDTM_issues", cfg.reports,
                  head_sum[_rpt_mask("sdtm_report")],
                  _rpt_det(open_iss, "sdtm_report"), "SDTM", run_label)
    _write_report("ADAM_issues", cfg.reports,
                  head_sum[_rpt_mask("adam_report")],
                  _rpt_det(open_iss, "adam_report"), "ADAM", run_label)
    _write_report("all_open",    cfg.reports,
                  head_sum, open_iss,   "All Open",   run_label)
    _write_report("all_closed",  cfg.reports,
                  head_sum[head_sum["n_closed"] > 0],
                  closed_iss, "All Closed", run_label)

    logger.info(">> [reporter] All reports written to: %s", cfg.reports)
